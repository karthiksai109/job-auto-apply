"""
Agent 4: Excel Tracker
Maintains a comprehensive Excel spreadsheet of all job applications.
Columns: Job Title, Company, Platform, URL, Tech Stack, Status,
         Applied Date, Rounds, Last Updated, Notes
Auto-syncs from the central job database every N minutes.
"""
import os
from datetime import datetime
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, BarChart

from agents.config import EXCEL_TRACKER_PATH, JobStatus
from agents.job_database import get_all_jobs, get_stats
from agents.logger import get_logger

logger = get_logger("ExcelTrack")

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_COLORS = {
    JobStatus.SCRAPED: "D9E2F3",       # Light blue
    JobStatus.APPLYING: "FFF2CC",       # Light yellow
    JobStatus.APPLIED: "C6EFCE",        # Light green
    JobStatus.FAILED_TO_APPLY: "FFC7CE",# Light red
    JobStatus.SCREENING: "B4C6E7",      # Blue
    JobStatus.PHONE_SCREEN: "9BC2E6",   # Medium blue
    JobStatus.TECHNICAL_ROUND: "8DB4E2", # Darker blue
    JobStatus.ONSITE: "5B9BD5",         # Even darker blue
    JobStatus.FINAL_ROUND: "4472C4",    # Dark blue
    JobStatus.OFFER: "00B050",          # Green
    JobStatus.REJECTED: "FF0000",       # Red
    JobStatus.WITHDRAWN: "808080",      # Gray
    JobStatus.NO_RESPONSE: "FFC000",    # Orange
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column definitions
COLUMNS = [
    ("Job Title", 35),
    ("Company", 25),
    ("Platform", 12),
    ("Location", 20),
    ("URL", 45),
    ("Tech Stack", 40),
    ("Status", 18),
    ("Applied Date", 18),
    ("Scraped Date", 18),
    ("Last Updated", 18),
    ("Rounds", 35),
    ("Search Query", 20),
    ("Notes", 30),
]


class ExcelTrackerAgent:
    """
    Agent 4: Maintains a live Excel spreadsheet tracking all job applications.
    - Creates/updates the Excel file from the central database
    - Color-codes by status
    - Includes summary dashboard sheet with charts
    - Auto-formats for easy reading
    """

    def __init__(self):
        self.excel_path = EXCEL_TRACKER_PATH
        self.jobs_synced = 0

    def run(self) -> dict:
        """Sync all jobs to Excel and generate dashboard."""
        logger.info("=" * 60)
        logger.info("AGENT 4: Excel Tracker Syncing...")
        logger.info("=" * 60)

        all_jobs = get_all_jobs()
        logger.info(f"Total jobs in database: {len(all_jobs)}")

        # Create/update the Excel workbook
        wb = self._create_workbook()

        # Write jobs to "All Applications" sheet
        ws_all = wb["All Applications"]
        self._write_jobs(ws_all, all_jobs)

        # Create platform-specific sheets
        platforms = set(j.get("platform", "unknown") for j in all_jobs)
        for platform in sorted(platforms):
            sheet_name = platform.title()[:31]  # Excel sheet name limit
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws_platform = wb[sheet_name]
            platform_jobs = [j for j in all_jobs if j.get("platform") == platform]
            self._write_jobs(ws_platform, platform_jobs)
            logger.info(f"  {platform.title()}: {len(platform_jobs)} jobs")

        # Create Dashboard sheet
        self._create_dashboard(wb, all_jobs)

        # Save
        wb.save(self.excel_path)
        self.jobs_synced = len(all_jobs)
        logger.info(f"\nExcel saved: {self.excel_path}")
        logger.info(f"Synced {self.jobs_synced} jobs across {len(platforms)} platforms")

        return {
            "file": self.excel_path,
            "total_jobs": len(all_jobs),
            "platforms": list(platforms),
        }

    def _create_workbook(self) -> openpyxl.Workbook:
        """Create a new workbook or load existing one."""
        wb = openpyxl.Workbook()
        # Rename default sheet
        ws = wb.active
        ws.title = "All Applications"
        return wb

    def _write_jobs(self, ws, jobs: list):
        """Write job data to a worksheet with formatting."""
        # Clear existing data
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(COLUMNS)):
            for cell in row:
                cell.value = None

        # Write headers
        for col_idx, (header, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        # Write data rows
        for row_idx, job in enumerate(jobs, 2):
            tech_stack = job.get("tech_stack", [])
            if isinstance(tech_stack, list):
                tech_str = ", ".join(tech_stack)
            else:
                tech_str = str(tech_stack)

            rounds = job.get("rounds", [])
            if isinstance(rounds, list):
                rounds_str = " | ".join(
                    f"{r.get('type', 'unknown')}: {r.get('details', '')}" for r in rounds
                ) if rounds else ""
            else:
                rounds_str = str(rounds)

            row_data = [
                job.get("title", ""),
                job.get("company", ""),
                job.get("platform", "").title(),
                job.get("location", ""),
                job.get("url", ""),
                tech_str,
                job.get("status", ""),
                job.get("applied_at", "")[:10] if job.get("applied_at") else "",
                job.get("scraped_at", "")[:10] if job.get("scraped_at") else "",
                job.get("last_checked", "")[:10] if job.get("last_checked") else "",
                rounds_str,
                job.get("search_query", ""),
                job.get("notes", ""),
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # Make URLs clickable
                if col_idx == 5 and value:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

            # Color-code the status column (column 7)
            status = job.get("status", "")
            if status in STATUS_COLORS:
                ws.cell(row=row_idx, column=7).fill = PatternFill(
                    start_color=STATUS_COLORS[status],
                    end_color=STATUS_COLORS[status],
                    fill_type="solid",
                )

        # Set row height
        for row_idx in range(1, len(jobs) + 2):
            ws.row_dimensions[row_idx].height = 22

    def _create_dashboard(self, wb: openpyxl.Workbook, all_jobs: list):
        """Create a summary dashboard sheet with stats and charts."""
        if "Dashboard" in wb.sheetnames:
            del wb["Dashboard"]

        ws = wb.create_sheet("Dashboard", 0)  # First sheet

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="Job Application Dashboard")
        title_cell.font = Font(name="Calibri", bold=True, size=18, color="2F5496")
        title_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=1, value=f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

        # Summary stats
        stats = get_stats()
        row = 4

        ws.cell(row=row, column=1, value="SUMMARY STATISTICS").font = Font(bold=True, size=14, color="2F5496")
        row += 1

        summary_items = [
            ("Total Jobs Tracked", stats["total"]),
            ("Applied Today", stats["today_applied"]),
        ]
        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Status breakdown
        ws.cell(row=row, column=1, value="BY STATUS").font = Font(bold=True, size=14, color="2F5496")
        row += 1

        ws.cell(row=row, column=1, value="Status").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Count").font = Font(bold=True)
        row += 1

        status_start_row = row
        for status in JobStatus.ALL:
            count = stats["by_status"].get(status, 0)
            ws.cell(row=row, column=1, value=status.replace("_", " ").title())
            ws.cell(row=row, column=2, value=count)
            if status in STATUS_COLORS:
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color=STATUS_COLORS[status],
                    end_color=STATUS_COLORS[status],
                    fill_type="solid",
                )
            row += 1

        # Platform breakdown
        row += 1
        ws.cell(row=row, column=1, value="BY PLATFORM").font = Font(bold=True, size=14, color="2F5496")
        row += 1

        ws.cell(row=row, column=1, value="Platform").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Count").font = Font(bold=True)
        row += 1

        for platform, count in sorted(stats["by_platform"].items()):
            ws.cell(row=row, column=1, value=platform.title())
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Tech stack breakdown
        row += 1
        ws.cell(row=row, column=1, value="TOP TECH STACKS").font = Font(bold=True, size=14, color="2F5496")
        row += 1

        tech_count = {}
        for job in all_jobs:
            for tech in job.get("tech_stack", []):
                tech_count[tech] = tech_count.get(tech, 0) + 1

        ws.cell(row=row, column=1, value="Technology").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Job Count").font = Font(bold=True)
        row += 1

        for tech, count in sorted(tech_count.items(), key=lambda x: -x[1])[:20]:
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        # Try to add a pie chart for status distribution
        try:
            if len(stats["by_status"]) > 0:
                chart = PieChart()
                chart.title = "Applications by Status"
                chart.style = 10
                data = Reference(ws, min_col=2, min_row=status_start_row - 1,
                                max_row=status_start_row + len(JobStatus.ALL) - 1)
                cats = Reference(ws, min_col=1, min_row=status_start_row,
                                max_row=status_start_row + len(JobStatus.ALL) - 1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.width = 18
                chart.height = 12
                ws.add_chart(chart, "D4")
        except Exception as e:
            logger.debug(f"Chart creation error: {e}")
